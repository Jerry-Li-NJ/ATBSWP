#! /usr/bin/python3

# SSToTest.py -- Open a spreadsheet and write the contents on columns to text files seperately.

import openpyxl, sys, os
from openpyxl.utils import get_column_letter

# Get spreadsheet name from command line.
if len(sys.argv) != 2:
	print('''
	Watch!!! The usage should be like this: (Only one parameter allowed)
		./SSToText.py test_xxx.xlxs
		''')
elif not sys.argv[1].endswith('.xlsx'):
	print('''
	Watch!!! The usage should be like this: (parameter should be a xlsx file)
		./SSToText.py test_xxx.xlxs
		''')
else:
	#print('test')

# Load the spreadsheet.
	wb = openpyxl.load_workbook(sys.argv[1])
	sheet = wb.active

# Read in contents and write to files in a loop.
	fileList = []
	for column in range(1, sheet.max_column + 1):
		filename = 'test_' + str(column)
		# Detele file, if it already exists
		if os.path.exists(filename):
			os.remove(filename)
		file = open(filename, 'a')
		fileList.append(filename)	# Store the filename to a list 
		for row in range(1, sheet.max_row + 1):	
			value = sheet.cell(row=row, column=column).value
			#print(value)
			if value == None:
				file.write('\n')
			else:
				file.write(value + '\n')
		file.close()



		print('\n')
		print('\tWrite done, cloumn' + get_column_letter(column) + 'write to file ' +filename)
	

	# Delete the blank lines at the end of the file
	for fileName in fileList:
		newFile = open(fileName)
		contentList = newFile.readlines()
		# Remove the blank lines in the list
		for i in range(len(contentList), 0, -1):
			if contentList[i-1] != '\n':
				break
			else:
				contentList.pop(i-1)
		newFile.close()
		print(contentList)
		# Write the first line and overwrite the file 
		file = open(fileName, 'w')
		value = contentList[0]
		file.write(value)
		file.close()
		# Write the last lines
		file = open(fileName, 'a')
		for i in range(1, len(contentList)):	
			value = contentList[i]
			file.write(value)
		file.close()


