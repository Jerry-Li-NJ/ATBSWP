#! /usr/bin/python3

# multiplocationTable.py - generate a N x N table in a spreadsheet with a number take from command line.

import openpyxl, sys
from openpyxl.styles import Font

'''
def CreateAWorkbook():
	wb = openpyxl.Workbook()
	sheet = wb.active

	return wb
'''

# Get number N from command.
## Make sure the input data type and number of input argument.
if len(sys.argv) != 2:
	print('''
		Wrong format!!!
		Use below format:
			./multiplocationTable.py 6
		''')
else:
	try:
		N = sys.argv[1]
		N = int(sys.argv[1])
	
	except ValueError:
		print('''
			Please use an integear as an parameter!!!
			''')
	if type(N) != type(1):
		pass
	elif N <0 or N == 0:
		print('''
			Please use an integear larger than 0!!!
			''')
	else:
# Create a workbook.
		#realWb = CreateAWorkbook()
		wb = openpyxl.Workbook()
		sheet = wb.active

# Fill in all labels on column A and row 1, and set them to bold.
		for i in range(2, N+2):
			cellObj1 = sheet.cell(row=1, column=i, value=i-1)
			cellObj2 = sheet.cell(row=i, column=1, value=i-1)
			cellObj1.font = Font(bold=True)
			cellObj2.font = Font(bold=True)

# Fill in the last integers.
		for i in range(2, N+2):
			for j in range(2, N+2):
				cellObj1 = sheet.cell(row=i, column=j, value=(i-1)*(j-1))

# Save the workbook to a xlsx file.
		wb.save('MT.xlsx')
