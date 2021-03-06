#! /usr/bin/python3
# updateProduce.py - Correct costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PROCE_UPDATES = {'Garlic': 3.07,
		'Celery': 1.19,
		'Lemon': 1.27}

# Loop through the rows and update the price.
for rowNum in range(2, sheet.max_row):	# skip the first row
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PROCE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PROCE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')


