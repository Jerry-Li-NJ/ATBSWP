#! /usr/bin/python3
# Chapter14_ETCC - Reads all Excel Files in current working diectory, and outpus them as CSV files.

from openpyxl import load_workbook
import os, csv

# Loop over all the Excel files.
 
'''
First, we need to loop over all the Excels.
Only Excel, so we need to skip all the non-xlsx files.
For every Excel, we should load it first, then we can get the data inside.
'''

i = 0
for excelFile in os.listdir('.'):
	if not excelFile.endswith('xlsx'):
		continue			# Skip non-xlsx files
	else:
		i = i + 1			# Use to count the amount of Excel files
		#print(excelFile + str(i))	# Uncomment to print the count for Excels
		print('### Load the Excel ' + excelFile + '...')
		wb = load_workbook(excelFile)	# Load the workbook object

# Loop over sheets in current Excel file.

		'''
		Create the name for the csv file.
		Create the File object, then the csv.writer object, just as the flow.
		'''

		j = 0
		for sheet in wb:
			j = j + 1			# use to count the amount of sheets
			#print('\t' + sheet.title + '  ' + str(j))	# Uncomment to show the sheets
			csvName = excelFile.split('.')[0] + '_' + sheet.title + '.csv'	# Create the CSV filename from the Excel filename and sheet title
			print('Creating the csv file: ' + csvName + '...')
			fileObj = open(csvName, 'w', newline = '')
			writerObj = csv.writer(fileObj)			# Create the csv.writer object for this CSV file

# Loop over rows in current sheet.

			'''
			Loop over all the rows to get all the cell.
			Actually, it's an embed loop.
			Why save all cells in a row to a list? 'cause there is a writerow method, that only accept list.
			'''

			print('Writeing data to csv file...')
			for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row = sheet.max_row):
				rowData = []		# Append each cell to this list
				for cell in row:	# Loop through each cell in the row
					rowData.append(cell.value)

# Write current row to the CSV file.

				'''
				Eventually we write the cells in a row to the csv file.
				After the loop done, all data inside the sheet is writen to the csv file.
				'''

				writerObj.writerow(rowData)
			print('Writeing data to csv file done.\n\n')

			fileObj.close()
	
