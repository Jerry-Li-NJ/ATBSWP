#! /usr/bin/python3

# SSCellInverter.py -- Invert the row and column of cells in the spreedsheet.

import openpyxl, sys, pprint

# Get the spreadsheet name from command line.
if len(sys.argv) != 2:
	print('''
	Please make sure one parameter and only one!!!
	Like this:
		./SSCellInverter.py aaa.xlsx
	
	''')
elif 'xlsx' not in sys.argv[1]:
	print('''
	Please make sure the parameter you pass is a spreadsheet name.
	Like this:
		./SSCellInverter.py aaa.xlsx
	
	''')
else:
	

# Load the spreadsheet.
	SSName = sys.argv[1]
	print(SSName)
	wb = openpyxl.load_workbook(SSName)
	sheet = wb.active

# Get the data inside and store to nested list.
	StoreList = []
	print('Storing data ............')
	for i in range(1, sheet.max_row+1):
		StoreList.append([])
		for j in range(1, sheet.max_column+1):
			cellValue = sheet.cell(row=i, column=j).value
			StoreList[i-1].append(cellValue)
	pprint.pprint(StoreList)

# Create a new workbook.
	newWB = openpyxl.Workbook()
	newSheet = newWB.active

# Apply the invension to the new workbook.
	print('Writing data to new spreadsheet...........')
	for i in range(1, sheet.max_row+1):
		for j in range(1, sheet.max_column+1):
			newSheet.cell(row=j, column=i, value=StoreList[i-1][j-1])

# Save the workbook to a new spreadsheet.
	updatedSSNameList = SSName.split('.')
	updatedSSName = updatedSSNameList[0] + '_updated.' + updatedSSNameList[1]
	newWB.save(updatedSSName)
	
	print('Done!')


