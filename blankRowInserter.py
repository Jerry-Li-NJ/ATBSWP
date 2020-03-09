#! /usr/bin/python3

# blankRowInserter.py -- Insert M blank lines at line N in a xlsx file.

import sys, openpyxl, pprint, shelve, sys
from openpyxl.styles import Font

#sys.setrecursionlimit(1000000)
def write_data(newWb, start, end, sheetObj, DS, step):	
	# pass new WrokBook obj, start and end row number, original sheet, DS, and the step, then you get the one you want -- an new WB obj.
	# newWB -- the new updated WB obj
	# start -- the start row
	# end -- the end row + 1
	# sheetObj -- the worksheet obj of the original SS
	# DS -- data stored
	# step -- for the blank rows


	newSheet = newWb.active

	for i in range(start, end):
		for j in range(1, sheetObj.max_column+1):
			newI = i-step
			newCell = newSheet.cell(row=i, column=j, value=storeDS[newI]['data'][str(j)])

			newName = DS[newI]['font'][str(j)]['name']
			newSize = DS[newI]['font'][str(j)]['size']
			newBold = DS[newI]['font'][str(j)]['bold']
			newItalic = DS[newI]['font'][str(j)]['italic']
			newVertAlign = DS[newI]['font'][str(j)]['vertAlign']
			newUnderline = DS[newI]['font'][str(j)]['underline']
			newStrike = DS[newI]['font'][str(j)]['strike']
			newColor = DS[newI]['font'][str(j)]['color']

			newFont = Font(name=newName,
					size=newSize,
					bold=newBold,
					italic=newItalic,
					vertAlign=newVertAlign,
					underline=newUnderline,
					strike=newStrike,
					color=newColor)
			newCell.font = newFont
			
	return newWb

# Get N, M and filename from command.
if len(sys.argv) != 4:	# Check the length
	print('''
		Wrong format!!!
		Use below format:
			./blankRowInserter.py 3 2 myProduce.xlsx
		''')
else:
	try:		# Make sure N&M are integers
		N = sys.argv[1]
		N = int(sys.argv[1])
		
		M = sys.argv[2]
		M = int(sys.argv[2])

	except ValueError:
		print('''
			Please use an integear as an parameter for N&M!!!
			''')
	if type(N) != type(1) or type(M) != type(1):
		pass
	elif N <0 or N == 0 or M <0 or M == 0:	# Make sure N&M are valid integers
		print('''
			Please use an integear larger than 0 for N&M!!!
			''')
	elif 'xlsx' not in sys.argv[3]:		# Make sure filename correct
		print('''
			Please use correct filename(with xlsx extension)!!!
			''')
	else:
		#print('hello')

# Read the spreadsheet, then store that data and font to a data structure.
		spreadsheet = sys.argv[3]
		wb = openpyxl.load_workbook(spreadsheet)
		sheet = wb.active
		
		storeDS = {}
		for i in range(1, sheet.max_row+1):
			storeDS[i] = {}
			storeDS[i]['data'] = {}
			storeDS[i]['font'] = {}

			#pprint.pprint(storeDS)
			for j in range(1, sheet.max_column+1):
				theCell = sheet.cell(row=i, column=j)

				storeDS[i]['data'][str(j)] = theCell.value
				storeDS[i]['font'][str(j)] = {}

				# Store the arguments of the cell objeck's font
				storeDS[i]['font'][str(j)]['name'] = theCell.font.name
				storeDS[i]['font'][str(j)]['size'] = theCell.font.size
				storeDS[i]['font'][str(j)]['bold'] = theCell.font.bold
				storeDS[i]['font'][str(j)]['italic'] = theCell.font.italic
				storeDS[i]['font'][str(j)]['vertAlign'] = theCell.font.vertAlign
				storeDS[i]['font'][str(j)]['underline'] = theCell.font.underline
				storeDS[i]['font'][str(j)]['strike'] = theCell.font.strike
				storeDS[i]['font'][str(j)]['color'] = theCell.font.color
		#pprint.pprint(storeDS)

# Create a new wb.
		newCreatedWb = openpyxl.Workbook()
		#newSheet = newWb.active
		
# Write first N -1 lines.
		generatedWb = write_data(newCreatedWb, 1, N, sheet, storeDS, 0)

# Writing the M blank lines.
# Write the last lines.
		generatedWbv2 = write_data(newCreatedWb, N+M, sheet.max_row+1+M, sheet, storeDS, M)

# Save the wb to a new file.
		SSList = spreadsheet.split('.')
		updatedSpreadsheet = SSList[0] + '_updated.' + SSList[1]
		generatedWbv2.save(updatedSpreadsheet)
